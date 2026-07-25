import logging
import os
import io
from datetime import datetime
from io import BytesIO

import joblib
import matplotlib.pyplot as plt
from celery import shared_task
from django.conf import settings
from django.core.files import File
from django.db.models import Count
from django.db.models.functions import TruncMonth
from django.utils import timezone
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from .models import Report
from equipment.models import Equipment
from maintenance.models import Intervention
from intelligence.models import SensorData
from notifications.tasks import send_alert_email

logger = logging.getLogger(__name__)


@shared_task
def generate_report_task(user_id, report_type, format_type, filters):
    logger.info(f"🔍 Génération rapport - user_id: {user_id}, type: {report_type}, format: {format_type}")

    from django.contrib.auth import get_user_model
    User = get_user_model()
    try:
        user = User.objects.get(id=user_id)
        logger.info(f"✅ Utilisateur trouvé : {user.username} (ID: {user.id})")
    except User.DoesNotExist:
        logger.error(f"❌ Utilisateur {user_id} introuvable !")
        user = None

    if format_type == 'pdf':
        file_content = generate_pdf_report(report_type, filters)
        extension = 'pdf'
    else:
        file_content = generate_excel_report(report_type, filters)
        extension = 'xlsx'

    file_name = f"report_{report_type}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.{extension}"
    report = Report.objects.create(
        title=f"Rapport {report_type} - {timezone.now().date()}",
        type=report_type,
        format=format_type,
        generated_by=user,
        parameters=filters,
    )
    report.file.save(file_name, File(BytesIO(file_content)), save=True)

    if user and user.email:
        send_alert_email.delay(
            subject=f"Rapport {report_type} généré",
            message=f"Votre rapport {report_type} est disponible dans l'application.",
            recipient_list=[user.email]
        )

    return report.id


def generate_pdf_report(report_type, filters):
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.HexColor('#1e3a8a'),
        spaceAfter=30,
        alignment=1,
    )
    heading_style = ParagraphStyle(
        'CustomHeading',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.HexColor('#1e293b'),
        spaceAfter=12,
    )
    normal_style = styles['Normal']

    elements = []

    # --- Titre ---
    title = f"Rapport {dict(Report.TYPE_CHOICES).get(report_type, report_type).upper()}"
    elements.append(Paragraph(title, title_style))
    now = timezone.now().strftime('%d/%m/%Y à %H:%M')
    elements.append(Paragraph(f"Généré le {now}", normal_style))
    elements.append(Spacer(1, 0.5*cm))

    # Filtres
    if filters:
        filter_text = "Filtres appliqués : "
        if filters.get('date_from'):
            filter_text += f"Du {filters['date_from']} "
        if filters.get('date_to'):
            filter_text += f"au {filters['date_to']} "
        if filters.get('machine'):
            try:
                machine = Equipment.objects.get(id=filters['machine'])
                filter_text += f"| Machine : {machine.name}"
            except Equipment.DoesNotExist:
                filter_text += f"| Machine ID: {filters['machine']}"
        elements.append(Paragraph(filter_text, normal_style))
    elements.append(Spacer(1, 0.5*cm))

    # --- Statistiques ---
    total_machines = Equipment.objects.count()
    active_machines = Equipment.objects.filter(status='active').count()
    maintenance_machines = Equipment.objects.filter(status='maintenance').count()
    out_machines = Equipment.objects.filter(status='out_of_service').count()

    stats_data = [
        ['Total machines', str(total_machines)],
        ['Actives', str(active_machines)],
        ['En maintenance', str(maintenance_machines)],
        ['Hors service', str(out_machines)],
    ]
    stats_table = Table(stats_data, colWidths=[6*cm, 3*cm])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#e2e8f0')),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#1e293b')),
        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
    ]))
    elements.append(Paragraph("📊 Résumé du parc", heading_style))
    elements.append(stats_table)
    elements.append(Spacer(1, 0.5*cm))

    # --- Graphique des interventions mensuelles ---
    elements.append(Paragraph("📈 Évolution mensuelle des interventions", heading_style))
    six_months_ago = timezone.now() - timezone.timedelta(days=180)
    monthly_data = (
        Intervention.objects
        .filter(created_at__gte=six_months_ago)
        .annotate(month=TruncMonth('created_at'))
        .values('month')
        .annotate(count=Count('id'))
        .order_by('month')
    )
    if monthly_data:
        months = [item['month'].strftime('%Y-%m') for item in monthly_data]
        counts = [item['count'] for item in monthly_data]
        plt.figure(figsize=(6, 3))
        plt.bar(months, counts, color='#2563eb')
        plt.title('Interventions par mois')
        plt.xlabel('Mois')
        plt.ylabel('Nombre')
        plt.xticks(rotation=45)
        plt.tight_layout()
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=100)
        buf.seek(0)
        # ✅ On passe directement buf à Image (plus besoin de ImageReader)
        elements.append(Image(buf, width=12*cm, height=6*cm))
        plt.close()
    else:
        elements.append(Paragraph("Aucune donnée d'intervention récente.", normal_style))
    elements.append(Spacer(1, 0.5*cm))

    # --- Tableau des équipements avec score de risque ---
    equipments = Equipment.objects.all()
    if filters and filters.get('machine'):
        equipments = equipments.filter(id=filters['machine'])

    model_path = os.path.join(settings.BASE_DIR, 'intelligence', 'models', 'random_forest.pkl')
    model = None
    if os.path.exists(model_path):
        model = joblib.load(model_path)
        from intelligence.tasks import extract_features

    table_data = [['Nom', 'Modèle', 'Statut', 'Criticité', 'Localisation', 'Risque (%)']]
    for eq in equipments:
        risk = '—'
        if model:
            latest = SensorData.objects.filter(machine=eq).order_by('-timestamp')[:24]
            if len(latest) >= 24:
                features = extract_features(latest)
                proba = model.predict_proba(features)[0][1]
                risk = f"{round(proba * 100, 1)}%"
        status_map = {'active': 'Actif', 'maintenance': 'En maintenance', 'out_of_service': 'Hors service'}
        criticality_map = {'low': 'Basse', 'medium': 'Moyenne', 'high': 'Élevée'}
        table_data.append([
            eq.name,
            eq.model,
            status_map.get(eq.status, eq.status),
            criticality_map.get(eq.criticality, eq.criticality),
            eq.location or '—',
            risk
        ])

    col_widths = [3.5*cm, 3.5*cm, 3*cm, 2.5*cm, 3.5*cm, 2.5*cm]
    eq_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    eq_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.HexColor('#ffffff')]),
    ]))
    elements.append(Paragraph("📋 Liste des équipements", heading_style))
    elements.append(eq_table)
    elements.append(Spacer(1, 0.5*cm))

    # --- Interventions récentes ---
    thirty_days_ago = timezone.now() - timezone.timedelta(days=30)
    recent_interventions = Intervention.objects.filter(created_at__gte=thirty_days_ago)[:10]
    if recent_interventions:
        elements.append(Paragraph("🔧 Interventions récentes (30 jours)", heading_style))
        inter_data = [['Machine', 'Type', 'Statut', 'Date']]
        for inv in recent_interventions:
            inter_data.append([
                inv.machine.name if inv.machine else 'N/A',
                inv.type,
                inv.status,
                inv.created_at.strftime('%d/%m/%Y')
            ])
        inter_table = Table(inter_data, colWidths=[4*cm, 3*cm, 3.5*cm, 3.5*cm])
        inter_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1e3a8a')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#cbd5e1')),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#f8fafc')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.HexColor('#f8fafc'), colors.HexColor('#ffffff')]),
        ]))
        elements.append(inter_table)

    elements.append(Spacer(1, 1*cm))
    elements.append(Paragraph(
        f"Document généré par OCP Maintenance - {datetime.now().strftime('%Y')}",
        ParagraphStyle('Footer', parent=normal_style, fontSize=8, textColor=colors.HexColor('#94a3b8'), alignment=1)
    ))

    doc.build(elements)
    buffer.seek(0)
    return buffer.getvalue()


def generate_excel_report(report_type, filters):
    wb = Workbook()

    ws = wb.active
    ws.title = "Équipements"
    ws.append(["Nom", "Modèle", "Statut", "Criticité", "Localisation"])
    equipments = Equipment.objects.all()
    if filters and filters.get('machine'):
        equipments = equipments.filter(id=filters['machine'])
    for machine in equipments:
        ws.append([
            machine.name,
            machine.model,
            machine.status,
            machine.criticality,
            machine.location or ''
        ])

    ws_stats = wb.create_sheet("Statistiques")
    ws_stats.append(["Métrique", "Valeur"])
    ws_stats.append(["Total machines", Equipment.objects.count()])
    ws_stats.append(["Actives", Equipment.objects.filter(status='active').count()])
    ws_stats.append(["En maintenance", Equipment.objects.filter(status='maintenance').count()])
    ws_stats.append(["Hors service", Equipment.objects.filter(status='out_of_service').count()])

    ws_risk = wb.create_sheet("Risques")
    ws_risk.append(["Machine", "Dernier score (%)", "Date du calcul"])
    model_path = os.path.join(settings.BASE_DIR, 'intelligence', 'models', 'random_forest.pkl')
    if os.path.exists(model_path):
        model = joblib.load(model_path)
        from intelligence.tasks import extract_features
        for machine in Equipment.objects.all():
            latest = SensorData.objects.filter(machine=machine).order_by('-timestamp')[:24]
            if len(latest) >= 24:
                features = extract_features(latest)
                proba = model.predict_proba(features)[0][1]
                risk = round(proba * 100, 1)
                last_ts = latest[0].timestamp.strftime('%Y-%m-%d %H:%M')
                ws_risk.append([machine.name, risk, last_ts])
            else:
                ws_risk.append([machine.name, '—', 'Données insuffisantes'])

    ws_inter = wb.create_sheet("Interventions")
    ws_inter.append(["Machine", "Type", "Statut", "Date"])
    interventions = Intervention.objects.all().order_by('-created_at')[:50]
    for inv in interventions:
        ws_inter.append([
            inv.machine.name if inv.machine else '',
            inv.type,
            inv.status,
            inv.created_at.strftime('%d/%m/%Y')
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()